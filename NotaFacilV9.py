import streamlit as st
import pandas as pd
import sqlite3
import os
from io import BytesIO

# Configuração visual e layout da página
st.set_page_config(page_title="Controle Escolar - NotaFácil", page_icon="🎓", layout="wide")

# --- INJEÇÃO DE CSS PERSONALIZADO (CANTOS ARREDONDADOS E CORES SUAVES/ALEGRES) --
st.markdown("""
<style>
    /* Estilos globais e fontes */
    html, body, [data-testid="stAppViewContainer"] {
        background-color: #f7f9fc;
        font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
    }
    
    /* Configurações da barra lateral */
    [data-testid="stSidebar"] {
        background-color: #ffffff !important;
        border-right: 1px solid #e2e8f0;
        box-shadow: 2px 0 10px rgba(0, 0, 0, 0.02);
    }
    
    /* Cantos arredondados gerais para inputs e caixas */
    .stTextInput > div > div > input, 
    .stSelectbox > div > div > div, 
    .stNumberInput > div > div > input {
        border-radius: 12px !important;
        border: 1.5px solid #e2e8f0 !important;
        padding: 4px 10px !important;
        transition: all 0.3s ease;
    }
    
    .stTextInput > div > div > input:focus, 
    .stSelectbox > div > div > div:focus, \n    .stNumberInput > div > div > input:focus {
        border-color: #6c5ce7 !important;
        box-shadow: 0 0 0 2px rgba(108, 92, 231, 0.1) !important;
    }
    
    /* Botões padrão alegres com cantos arredondados e cor roxo/lilás pastel */
    button[kind="primary"], .stForm button, div.stButton > button {
        border-radius: 15px !important;
        background: linear-gradient(135deg, #6c5ce7 0%, #a29bfe 100%) !important;
        color: white !important;
        border: none !important;
        padding: 8px 24px !important;
        font-weight: 600 !important;
        box-shadow: 0 4px 12px rgba(108, 92, 231, 0.2) !important;
        transition: all 0.3s ease !important;
    }
    
    button[kind="primary"]:hover, .stForm button:hover, div.stButton > button:hover {
        transform: translateY(-2px) !important;
        box-shadow: 0 6px 18px rgba(108, 92, 231, 0.35) !important;
        background: linear-gradient(135deg, #5b4bc4 0%, #8c7ae6 100%) !important;
    }

    /* Botões perigo (excluir/resetar) customizados através de wrapper classe .danger-button */
    .danger-button button, .danger-button div.stButton > button, .danger-button .stForm button {
        border-radius: 15px !important;
        background: linear-gradient(135deg, #ff7675 0%, #fab1a0 100%) !important;
        color: white !important;
        border: none !important;
        padding: 8px 24px !important;
        font-weight: 600 !important;
        box-shadow: 0 4px 12px rgba(255, 118, 117, 0.2) !important;
        transition: all 0.3s ease !important;
    }
    .danger-button button:hover, .danger-button div.stButton > button:hover, .danger-button .stForm button:hover {
        transform: translateY(-2px) !important;
        box-shadow: 0 6px 18px rgba(255, 118, 117, 0.35) !important;
        background: linear-gradient(135deg, #e17055 0%, #ff7675 100%) !important;
    }
    
    /* Customizar container de formulários */
    [data-testid="stForm"] {
        border-radius: 20px !important;
        border: 1px solid #e2e8f0 !important;
        background-color: #ffffff !important;
        padding: 25px !important;
        box-shadow: 0 8px 16px rgba(0, 0, 0, 0.03) !important;
    }
    
    /* Estilização alegre para as métricas */
    [data-testid="stMetricValue"] {
        font-size: 2rem !important;
        font-weight: 700 !important;
        color: #6c5ce7 !important;
    }
    
    [data-testid="stMetricLabel"] {
        color: #4a5568 !important;
        font-weight: 500 !important;
    }
    
    /* Cartões informativos e mensagens de sucesso/erro */
    .stAlert {
        border-radius: 15px !important;
        box-shadow: 0 4px 6px rgba(0,0,0,0.02) !important;
    }
    
    /* Tabela / DataFrame arredondado */
    [data-testid="stDataFrame"] {
        border-radius: 15px !important;
        overflow: hidden !important;
        border: 1px solid #e2e8f0 !important;
    }
    
    /* Imagens arredondadas */
    img {
        border-radius: 15px !important;
    }
    
    /* Estilo de boletim individual */
    .boletim-card {
        background-color: #ffffff;
        padding: 20px;
        border-radius: 20px;
        border: 1px solid #e2e8f0;
        box-shadow: 0 8px 16px rgba(0,0,0,0.02);
        margin-bottom: 20px;
    }
</style>
""", unsafe_allow_html=True)

DB_FILE = "notafacil.db"

# --- FUNÇÕES DE BANCO DE DADOS (SQLITE) ---
def get_db_connection():
    conn = sqlite3.connect(DB_FILE, check_same_thread=False)
    return conn

def init_db():
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS alunos (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        nome TEXT NOT NULL,
        turma TEXT NOT NULL,
        nota1 REAL,
        nota2 REAL,
        nota3 REAL,
        nota4 REAL,
        somatorio REAL,
        media REAL,
        status TEXT
    )
    """)
    
    # Verifica se deve inserir alunos padrão para demonstração inicial
    cursor.execute("SELECT COUNT(*) FROM alunos")
    if cursor.fetchone()[0] == 0:
        cursor.execute("""
        INSERT INTO alunos (nome, turma, nota1, nota2, nota3, nota4, somatorio, media, status)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, ("Maria Silva", "9º Ano", 8.5, 7.0, 9.0, 8.0, 32.5, 8.13, "Aprovado"))
        cursor.execute("""
        INSERT INTO alunos (nome, turma, nota1, nota2, nota3, nota4, somatorio, media, status)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, ("João Souza", "8º Ano", 5.0, 6.0, 0.0, 0.0, 11.0, 5.5, "Reprovado"))
    conn.commit()
    conn.close()

# Executa inicialização do banco
init_db()

def get_alunos():
    conn = get_db_connection()
    df = pd.read_sql_query("SELECT * FROM alunos", conn)
    conn.close()
    return df

# --- LÓGICA DE MÉDIA DINÂMICA (APENAS NOTAS DIFERENTES DE ZERO) ---
def calcular_soma_media_status(n1, n2, n3, n4):
    notas = [n1, n2, n3, n4]
    # Considera apenas notas preenchidas e diferentes de zero
    notas_validas = [n for n in notas if n != 0.0 and n is not None]
    
    if not notas_validas:
        return 0.0, 0.0, "Reprovado"
    
    somatorio = sum(notas_validas)
    media = somatorio / len(notas_validas)
    status = "Aprovado" if media >= 7.0 else "Reprovado"
    return round(somatorio, 2), round(media, 2), status

# --- LÓGICA DO BOTÃO DESFAZER (UNDO) ---
def registrar_desfazer(tipo, dados):
    """
    Registra uma ação no session_state para permitir reversão.
    tipo: 'excluir', 'editar', ou 'tabela_lote'
    dados: dicionário ou lista contendo os dados antes da alteração
    """
    st.session_state.historico_desfazer = {
        "tipo": tipo,
        "dados": dados
    }

def desfazer_ultima_acao():
    if "historico_desfazer" in st.session_state and st.session_state.historico_desfazer is not None:
        tipo = st.session_state.historico_desfazer["tipo"]
        dados = st.session_state.historico_desfazer["dados"]
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        if tipo == "excluir":
            cursor.execute("""
                INSERT INTO alunos (nome, turma, nota1, nota2, nota3, nota4, somatorio, media, status)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (dados["nome"], dados["turma"], dados["nota1"], dados["nota2"], dados["nota3"], dados["nota4"], dados["somatorio"], dados["media"], dados["status"]))
            st.toast(f"↩️ Aluno '{dados['nome']}' foi restaurado com sucesso!")
            
        elif tipo == "editar":
            cursor.execute("""
                UPDATE alunos
                SET nome=?, turma=?, nota1=?, nota2=?, nota3=?, nota4=?, somatorio=?, media=?, status=?
                WHERE id=?
            """, (dados["nome"], dados["turma"], dados["nota1"], dados["nota2"], dados["nota3"], dados["nota4"], dados["somatorio"], dados["media"], dados["status"], dados["id"]))
            st.toast(f"↩️ Alterações de '{dados['nome']}' foram revertidas para o estado anterior!")
            
        elif tipo == "tabela_lote":
            # Limpa e restaura todos os registros do lote
            cursor.execute("DELETE FROM alunos")
            for aluno in dados:
                cursor.execute("""
                    INSERT INTO alunos (id, nome, turma, nota1, nota2, nota3, nota4, somatorio, media, status)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (aluno["id"], aluno["nome"], aluno["turma"], aluno["nota1"], aluno["nota2"], aluno["nota3"], aluno["nota4"], aluno["somatorio"], aluno["media"], aluno["status"]))
            st.toast("↩️ Todas as alterações da planilha em lote foram revertidas!")
            
        conn.commit()
        conn.close()
        
        # Limpa o histórico de desfazer após execução
        st.session_state.historico_desfazer = None
        st.rerun()

def exibir_banner_desfazer():
    if "historico_desfazer" in st.session_state and st.session_state.historico_desfazer is not None:
        tipo_acao = st.session_state.historico_desfazer["tipo"]
        dados_backup = st.session_state.historico_desfazer["dados"]
        
        if tipo_acao == "excluir":
            msg = f"↩️ Aluno **{dados_backup['nome']}** foi excluído do sistema."
        elif tipo_acao == "editar":
            msg = f"↩️ Os dados do aluno **{dados_backup['nome']}** foram modificados."
        elif tipo_acao == "tabela_lote":
            msg = "↩️ Várias notas foram editadas em lote na planilha."
            
        col_msg, col_und = st.columns([4, 1])
        with col_msg:
            st.info(f"{msg} Deseja desfazer a última alteração?")
        with col_und:
            if st.button("Desfazer Ação", key="btn_desfazer_global", use_container_width=True):
                desfazer_ultima_acao()

# --- FUNÇÃO DE CONVERSÃO PARA EXCEL ESTILIZADO (.XLSX) COM FALLBACK SEGURO PARA CSV ---
def converter_para_excel_estilizado(df):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Notas Língua Portuguesa')
            
            # Acessa a planilha interna para formatar
            worksheet = writer.sheets['Notas Língua Portuguesa']
            
            # Paleta de cores suaves (Lilás pastel profissional para combinar com o app da Rebeca)
            header_fill = PatternFill(start_color="FFA29BFE", end_color="FFA29BFE", fill_type="solid") 
            header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFFFF")
            cell_font = Font(name="Segoe UI", size=10)
            center_align = Alignment(horizontal="center", vertical="center")
            left_align = Alignment(horizontal="left", vertical="center")
            
            thin_border = Border(
                left=Side(style='thin', color='FFCBD5E1'),
                right=Side(style='thin', color='FFCBD5E1'),
                top=Side(style='thin', color='FFCBD5E1'),
                bottom=Side(style='thin', color='FFCBD5E1')
            )
            
            # Estilização do cabeçalho
            for col_num in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_align
                
            # Estilização de dados
            for row in range(2, worksheet.max_row + 1):
                for col in range(1, worksheet.max_column + 1):
                    cell = worksheet.cell(row=row, column=col)
                    cell.font = cell_font
                    cell.border = thin_border
                    
                    # Centraliza tudo que não é o nome do aluno
                    if col > 1:
                        cell.alignment = center_align
                    else:
                        cell.alignment = left_align
                        
            # Autoajuste automático de colunas tratando valores nulos e zeros
            for col in worksheet.columns:
                max_len = 0
                for cell in col:
                    val = cell.value
                    val_str = "" if val is None else str(val)
                    if len(val_str) > max_len:
                        max_len = len(val_str)
                col_letter = col[0].column_letter
                worksheet.column_dimensions[col_letter].width = max(max_len + 3, 12)
                
        processed_data = output.getvalue()
        return processed_data, True
    except Exception as e:
        # Fallback caso dê erro ou falte a biblioteca 'openpyxl' no Streamlit Cloud
        csv_data = df.to_csv(index=False).encode("utf-8")
        return csv_data, False

# Lista de turmas padrão do 1º ao 9º Ano
OPCOES_TURMAS = [f"{i}º Ano" for i in range(1, 10)]

# --- BARRA LATERAL (SIDEBAR) ---
st.sidebar.markdown("<div style='text-align: center;'>", unsafe_allow_html=True)
if os.path.exists("professora_loira.png"):
    st.sidebar.image("professora_loira.png", use_container_width=True)
else:
    st.sidebar.image("https://cdn-icons-png.flaticon.com/512/3413/3413535.png", width=120)

st.sidebar.markdown("<h3 style='text-align: center; color: #6c5ce7; margin-bottom: 5px; font-weight: bold;'>Prof.ª REBECA CARVALHO</h3>", unsafe_allow_html=True)
st.sidebar.markdown("<p style='text-align: center; font-style: italic; color: #a29bfe; margin-top: 0px;'>Língua Portuguesa & Literatura</p>", unsafe_allow_html=True)
st.sidebar.markdown("</div>", unsafe_allow_html=True)
st.sidebar.markdown("---")

st.sidebar.title("Menu de Navegação")
menu = st.sidebar.radio(
    "Selecione uma opção:",
    [
        "📋 Cadastrar Aluno", 
        "✏️ Editar Aluno (Formulário)",
        "📝 Edição Rápida (Planilha)",
        "🔍 Visualizar Aluno (Boletim)",
        "📊 Painel Geral (Resultados)"
    ]
)

# --- RENDEREZA O BANNER DE DESFAZER GLOBAL SE DISPONÍVEL ---
exibir_banner_desfazer()

# --- ABA 1: CADASTRAR ALUNO ---
if menu == "📋 Cadastrar Aluno":
    st.title("📋 Cadastrar Novo Aluno")
    st.markdown("""
    Insira o nome do aluno, selecione a turma (1º ao 9º Ano) e adicione as notas divididas de Língua Portuguesa.
    
    💡 **Lógica da Média Dinâmica:** O sistema calcula a média utilizando **apenas as notas preenchidas (maiores que zero)**. 
    Se você preencher apenas duas notas e deixar as outras como zero, a média será calculada dividindo por 2 (sem penalizar com zeros antes das provas serem aplicadas).
    """)

    with st.form(key="form_aluno", clear_on_submit=True):
        col_cad_1, col_cad_2 = st.columns(2)
        with col_cad_1:
            nome = st.text_input("Nome Completo do Aluno:", placeholder="Ex: Carlos Henrique de Oliveira")
        with col_cad_2:
            turma = st.selectbox("Turma (Ano Escolar):", OPCOES_TURMAS)
            
        st.markdown("### 📝 Notas de Avaliações")
        col_n1, col_n2 = st.columns(2)
        col_n3, col_n4 = st.columns(2)
        
        with col_n1:
            nota1 = st.number_input("Nota 1:", min_value=0.0, max_value=10.0, value=0.0, step=0.1, help="Deixe 0.0 se ainda não realizada")
        with col_n2:
            nota2 = st.number_input("Nota 2:", min_value=0.0, max_value=10.0, value=0.0, step=0.1, help="Deixe 0.0 se ainda não realizada")
        with col_n3:
            nota3 = st.number_input("Nota 3:", min_value=0.0, max_value=10.0, value=0.0, step=0.1, help="Deixe 0.0 se ainda não realizada")
        with col_n4:
            nota4 = st.number_input("Nota 4:", min_value=0.0, max_value=10.0, value=0.0, step=0.1, help="Deixe 0.0 se ainda não realizada")
        
        submeter_aluno = st.form_submit_button("💾 Salvar Aluno no Banco de Dados")
        
    if submeter_aluno:
        if not nome.strip():
            st.error("Por favor, preencha o nome do aluno!")
        else:
            soma, media, status = calcular_soma_media_status(nota1, nota2, nota3, nota4)
            
            # Salva no SQLite
            conn = get_db_connection()
            cursor = conn.cursor()
            cursor.execute("""
            INSERT INTO alunos (nome, turma, nota1, nota2, nota3, nota4, somatorio, media, status)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (nome.strip(), turma, nota1, nota2, nota3, nota4, soma, media, status))
            conn.commit()
            conn.close()
            
            st.success(f"🎉 Aluno **{nome}** cadastrado e salvo com sucesso!")
            st.rerun()

# --- ABA 2: EDITAR ALUNO (FORMULÁRIO) ---
elif menu == "✏️ Editar Aluno (Formulário)":
    st.title("✏️ Editar Aluno (Formulário Individual)")
    st.markdown("Selecione um aluno na caixa abaixo para carregar seus dados, efetuar alterações ou removê-lo.")
    
    df_alunos = get_alunos()
    
    if not df_alunos.empty:
        # Cria identificador único (nome + turma)
        df_alunos["Identificador"] = df_alunos["nome"] + " (" + df_alunos["turma"] + ")"
        identificadores = df_alunos["Identificador"].tolist()
        aluno_selecionado = st.selectbox("Escolha o aluno para carregar os dados:", identificadores)
        
        # Obtém o id correspondente do aluno
        aluno_dados = df_alunos[df_alunos["Identificador"] == aluno_selecionado].iloc[0]
        id_aluno = int(aluno_dados["id"])
        
        # Define índice inicial da turma na lista
        turma_atual = aluno_dados["turma"]
        index_turma = OPCOES_TURMAS.index(turma_atual) if turma_atual in OPCOES_TURMAS else 0
            
        with st.form(key="form_edicao_aluno"):
            col_ed_1, col_ed_2 = st.columns(2)
            with col_ed_1:
                novo_nome = st.text_input("Nome Completo:", value=aluno_dados["nome"])
            with col_ed_2:
                nova_turma = st.selectbox("Turma (Ano Escolar):", OPCOES_TURMAS, index=index_turma)
                
            st.markdown("### 📝 Notas de Avaliação")
            col_en1, col_en2 = st.columns(2)
            col_en3, col_en4 = st.columns(2)
            
            with col_en1:
                nova_nota1 = st.number_input("Nota 1:", min_value=0.0, max_value=10.0, value=float(aluno_dados["nota1"]), step=0.1)
            with col_en2:
                nova_nota2 = st.number_input("Nota 2:", min_value=0.0, max_value=10.0, value=float(aluno_dados["nota2"]), step=0.1)
            with col_en3:
                nova_nota3 = st.number_input("Nota 3:", min_value=0.0, max_value=10.0, value=float(aluno_dados["nota3"]), step=0.1)
            with col_en4:
                nova_nota4 = st.number_input("Nota 4:", min_value=0.0, max_value=10.0, value=float(aluno_dados["nota4"]), step=0.1)
                
            col_btns_1, col_btns_2 = st.columns(2)
            with col_btns_1:
                salvar_edicao = st.form_submit_button("💾 Salvar Alterações")
            with col_btns_2:
                st.markdown('<div class="danger-button">', unsafe_allow_html=True)
                excluir_aluno = st.form_submit_button("🗑️ Excluir Aluno")
                st.markdown('</div>', unsafe_allow_html=True)
                
        if salvar_edicao:
            if not novo_nome.strip():
                st.error("O nome não pode ficar vazio!")
            else:
                # Backup para o desfazer antes de alterar
                backup_edit = {
                    "id": id_aluno,
                    "nome": aluno_dados["nome"],
                    "turma": aluno_dados["turma"],
                    "nota1": aluno_dados["nota1"],
                    "nota2": aluno_dados["nota2"],
                    "nota3": aluno_dados["nota3"],
                    "nota4": aluno_dados["nota4"],
                    "somatorio": aluno_dados["somatorio"],
                    "media": aluno_dados["media"],
                    "status": aluno_dados["status"]
                }
                registrar_desfazer("editar", backup_edit)
                
                # Recalcula soma, média e situação (ignorando zeros)
                novo_somatorio, nova_media, nova_situacao = calcular_soma_media_status(nova_nota1, nova_nota2, nova_nota3, nova_nota4)
                
                # Salva no banco
                conn = get_db_connection()
                cursor = conn.cursor()
                cursor.execute("""
                    UPDATE alunos
                    SET nome=?, turma=?, nota1=?, nota2=?, nota3=?, nota4=?, somatorio=?, media=?, status=?
                    WHERE id=?
                """, (novo_nome.strip(), nova_turma, nova_nota1, nova_nota2, nova_nota3, nova_nota4, novo_somatorio, nova_media, nova_situacao, id_aluno))
                conn.commit()
                conn.close()
                
                st.success(f"🎉 Alterações para **{novo_nome}** salvas com sucesso no Banco de Dados!")
                st.rerun()
                
        if excluir_aluno:
            # Backup para o desfazer antes de deletar
            backup_del = {
                "nome": aluno_dados["nome"],
                "turma": aluno_dados["turma"],
                "nota1": aluno_dados["nota1"],
                "nota2": aluno_dados["nota2"],
                "nota3": aluno_dados["nota3"],
                "nota4": aluno_dados["nota4"],
                "somatorio": aluno_dados["somatorio"],
                "media": aluno_dados["media"],
                "status": aluno_dados["status"]
            }
            registrar_desfazer("excluir", backup_del)
            
            # Deleta do banco
            conn = get_db_connection()
            cursor = conn.cursor()
            cursor.execute("DELETE FROM alunos WHERE id=?", (id_aluno,))
            conn.commit()
            conn.close()
            st.success(f"🗑️ Aluno **{aluno_dados['nome']}** foi excluído com sucesso!")
            st.rerun()
            
    else:
        st.info("Nenhum aluno cadastrado no momento. Acesse '📋 Cadastrar Aluno' para registrar os dados.")

# --- ABA 3: EDIÇÃO RÁPIDA (PLANILHA) ---
elif menu == "📝 Edição Rápida (Planilha)":
    st.title("📝 Edição Rápida via Planilha")
    st.markdown("""
    Altere as notas ou dados de **múltiplos alunos de uma vez** direto na planilha abaixo!
    
    👉 **Instruções:**
    1. Dê um **duplo clique** na célula que deseja alterar (Nome, Turma, ou Notas).
    2. Modifique o valor.
    3. Clique no botão de **Salvar Alterações da Tabela** no final para recalcular e atualizar tudo simultaneamente.
    """)
    
    df_alunos = get_alunos()
    if not df_alunos.empty:
        # Prepara dataframe limpo para edição
        df_editor = df_alunos[["id", "nome", "turma", "nota1", "nota2", "nota3", "nota4", "somatorio", "media", "status"]].copy()
        df_editor.columns = ["ID", "Nome", "Turma", "Nota 1", "Nota 2", "Nota 3", "Nota 4", "Somatório (Automático)", "Média (Automática)", "Status"]
        
        edited_df = st.data_editor(
            df_editor,
            use_container_width=True,
            disabled=["ID", "Somatório (Automático)", "Média (Automática)", "Status"],
            column_config={
                "Turma": st.column_config.SelectboxColumn(
                    "Turma",
                    options=OPCOES_TURMAS,
                    required=True,
                )
            }
        )
        
        if st.button("💾 Salvar Alterações da Tabela"):
            # Registra backup da planilha completa para possibilitar desfazer
            backup_lote = df_alunos.to_dict(orient="records")
            registrar_desfazer("tabela_lote", backup_lote)
            
            conn = get_db_connection()
            cursor = conn.cursor()
            for _, row in edited_df.iterrows():
                try:
                    n1 = float(row["Nota 1"])
                    n2 = float(row["Nota 2"])
                    n3 = float(row["Nota 3"])
                    n4 = float(row["Nota 4"])
                except ValueError:
                    n1 = n2 = n3 = n4 = 0.0
                    
                soma, media, status = calcular_soma_media_status(n1, n2, n3, n4)
                
                cursor.execute("""
                    UPDATE alunos
                    SET nome=?, turma=?, nota1=?, nota2=?, nota3=?, nota4=?, somatorio=?, media=?, status=?
                    WHERE id=?
                """, (str(row["Nome"]).strip(), row["Turma"], n1, n2, n3, n4, soma, media, status, int(row["ID"])))
            conn.commit()
            conn.close()
            st.success("🎉 Todas as alterações feitas na planilha foram salvas com recálculo automático!")
            st.rerun()
    else:
        st.info("Nenhum aluno cadastrado no momento. Acesse '📋 Cadastrar Aluno' para registrar os dados.")

# --- ABA 4: VISUALIZAR ALUNO EM SEPARADO (BOLETIM COM FILTRO POR TURMAS) ---
elif menu == "🔍 Visualizar Aluno (Boletim)":
    st.title("🔍 Boletim Individual")
    st.markdown("Filtre pela turma do aluno para encontrar o boletim detalhado, médias e curvas de evolução.")
    
    df_alunos = get_alunos()
    if not df_alunos.empty:
        # Descobre as turmas reais cadastradas no banco de dados para o filtro
        turmas_cadastradas = sorted(df_alunos["turma"].unique().tolist())
        opcoes_filtro = ["Todas as Turmas"] + turmas_cadastradas
        
        col_filtro_1, col_filtro_2 = st.columns(2)
        with col_filtro_1:
            turma_selecionada = st.selectbox("🏫 Filtrar por Turma:", opcoes_filtro)
            
        # Filtra o dataframe com base na seleção da turma
        if turma_selecionada == "Todas as Turmas":
            df_filtrado = df_alunos.copy()
        else:
            df_filtrado = df_alunos[df_alunos["turma"] == turma_selecionada].copy()
            
        if not df_filtrado.empty:
            df_filtrado["Identificador"] = df_filtrado["nome"] + " (" + df_filtrado["turma"] + ")"
            lista_nomes = sorted(df_filtrado["Identificador"].unique().tolist())
            
            with col_filtro_2:
                aluno_selecionado = st.selectbox("👉 Escolha o aluno para ver o boletim:", lista_nomes)
            
            st.markdown("---")
            
            # Filtra os dados do aluno escolhido
            dados_aluno = df_filtrado[df_filtrado["Identificador"] == aluno_selecionado].iloc[0]
            
            # Boletim em cartão arredondado customizado
            st.markdown(f"""
            <div class="boletim-card">
                <h2 style='margin-top: 0px;'>Boletim Escolar de Língua Portuguesa</h2>\n                <hr style='margin-top: 5px; margin-bottom: 15px;'>\n            </div>
            """, unsafe_allow_html=True)
            
            col_b1, col_b2 = st.columns(2)
            
            with col_b1:
                st.markdown(f"## 👤 {dados_aluno['nome']}")
                st.markdown(f"### 🏫 Turma: **{dados_aluno['turma']}**")
                
                # Exibe status colorido
                if dados_aluno["status"] == "Aprovado":
                    st.success(f"🟢 **Status: Aprovado**")
                else:
                    st.error(f"🔴 **Status: Reprovado**")
                
                # Métricas grandes
                st.markdown("---")
                col_met1, col_met2 = st.columns(2)
                with col_met1:
                    st.metric(label="🎓 Média de Notas Ativas", value=f"{dados_aluno['media']:.2f}")
                with col_met2:
                    st.metric(label="➕ Somatório Total", value=f"{dados_aluno['somatorio']:.2f}")
                    
                st.info("💡 **Informação:** A média desconsidera as notas zeradas correspondentes a avaliações ainda não lançadas.")
                
            with col_b2:
                st.markdown("### 📊 Notas Detalhadas")
                
                # Notas
                col_n1_i, col_n2_i, col_n3_i, col_n4_i = st.columns(4)
                col_n1_i.metric("Nota 1", f"{dados_aluno['nota1']:.1f}")
                col_n2_i.metric("Nota 2", f"{dados_aluno['nota2']:.1f}")
                col_n3_i.metric("Nota 3", f"{dados_aluno['nota3']:.1f}")
                col_n4_i.metric("Nota 4", f"{dados_aluno['nota4']:.1f}")
                
                st.markdown("---")
                # Gráfico de evolução de desempenho
                st.markdown("**📉 Evolução por Avaliação:**")
                
                # Filtra apenas notas de interesse (maiores que zero para plotar evolução real)
                avaliacoes = []
                notas_plot = []
                if dados_aluno['nota1'] > 0:
                    avaliacoes.append("Nota 1")
                    notas_plot.append(dados_aluno['nota1'])
                if dados_aluno['nota2'] > 0:
                    avaliacoes.append("Nota 2")
                    notas_plot.append(dados_aluno['nota2'])
                if dados_aluno['nota3'] > 0:
                    avaliacoes.append("Nota 3")
                    notas_plot.append(dados_aluno['nota3'])
                if dados_aluno['nota4'] > 0:
                    avaliacoes.append("Nota 4")
                    notas_plot.append(dados_aluno['nota4'])
                    
                if avaliacoes:
                    df_temp = pd.DataFrame({
                        "Avaliação": avaliacoes,
                        "Nota": notas_plot
                    })
                    st.line_chart(data=df_temp, x="Avaliação", y="Nota")
                else:
                    st.warning("Nenhuma nota lançada para este aluno ainda.")
        else:
            st.warning("Nenhum aluno cadastrado nesta turma ainda.")
                
    else:
        st.info("Nenhum aluno cadastrado no momento. Acesse '📋 Cadastrar Aluno' para registrar os dados.")

# --- ABA 5: PAINEL GERAL (RESULTADOS) ---
elif menu == "📊 Painel Geral (Resultados)":
    st.title("📊 Painel Geral de Resultados")
    st.markdown("Consulte os dados consolidados da classe, médias globais de Língua Portuguesa e faça exportações no formato Excel ou CSV.")
    
    df_alunos = get_alunos()
    if not df_alunos.empty:
        # Estatísticas Globais
        total_alunos = len(df_alunos)
        media_geral = df_alunos["media"].mean()
        aprovados = len(df_alunos[df_alunos["status"] == "Aprovado"])
        
        col_st1, col_st2, col_st3 = st.columns(3)
        with col_st1:
            st.metric("Total de Alunos", total_alunos)
        with col_st2:
            st.metric("Média Geral Escolar", f"{media_geral:.2f}")
        with col_st3:
            st.metric("Aprovados", f"{aprovados} ({aprovados/total_alunos*100:.0f}%)")
        
        st.markdown("---")
        
        st.subheader("📋 Tabela Geral de Notas")
        # Renomeia colunas para visualização profissional
        df_exibicao = df_alunos[["nome", "turma", "nota1", "nota2", "nota3", "nota4", "somatorio", "media", "status"]].copy()
        df_exibicao.columns = ["Nome", "Turma", "Nota 1", "Nota 2", "Nota 3", "Nota 4", "Somatório", "Média", "Status"]
        st.dataframe(df_exibicao, use_container_width=True)
        
        # Gráfico Geral
        st.markdown("### 📈 Comparação Gráfica das Médias")
        st.bar_chart(data=df_exibicao, x="Nome", y="Média")
        
        # EXPORTAÇÃO EXCEL CUSTOMIZADA E CONTROLE COM FALLBACK SEGURO
        st.markdown("---")
        st.markdown("### 📥 Opções de Exportação Avançadas")
        st.markdown("Selecione o formato de exportação desejado para gerar uma planilha editável (Excel ou CSV):")
        
        tipo_exportacao = st.selectbox(
            "Selecione o escopo da exportação:",
            [
                "Planilha de Todos os Alunos (Toda a Escola)",
                "Planilha de uma Turma Específica",
                "Boletim Individual de um Aluno"
            ]
        )
        
        col_exp_1, col_exp_2 = st.columns(2)
        
        if tipo_exportacao == "Planilha de Todos os Alunos (Toda a Escola)":
            with col_exp_1:
                df_export = df_alunos[["nome", "turma", "nota1", "nota2", "nota3", "nota4", "somatorio", "media", "status"]].copy()
                df_export.columns = ["Nome do Aluno", "Turma", "Nota 1", "Nota 2", "Nota 3", "Nota 4", "Somatório", "Média", "Status"]
                
                export_data, is_excel = converter_para_excel_estilizado(df_export)
                
                if is_excel:
                    st.download_button(
                        label="📥 Baixar Planilha Consolidada Excel (.xlsx)",
                        data=export_data,
                        file_name="boletim_todos_alunos.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )
                else:
                    st.download_button(
                        label="📥 Baixar Planilha Consolidada CSV (.csv)",
                        data=export_data,
                        file_name="boletim_todos_alunos.csv",
                        mime="text/csv",
                        use_container_width=True
                    )
                    st.warning("⚠️ **Nota de Compatibilidade:** O pacote `openpyxl` não está instalado no seu ambiente Streamlit Cloud, por isso a planilha foi gerada automaticamente como **CSV** (totalmente editável e compatível com Excel). Para baixar o formato com design estilizado (.xlsx), inclua `openpyxl` no arquivo `requirements.txt` do seu repositório no GitHub!")
                    
        elif tipo_exportacao == "Planilha de uma Turma Específica":
            turmas_cadastradas = sorted(df_alunos["turma"].unique().tolist())
            if turmas_cadastradas:
                with col_exp_1:
                    turma_sel = st.selectbox("Selecione a Turma:", turmas_cadastradas)
                with col_exp_2:
                    df_turma = df_alunos[df_alunos["turma"] == turma_sel][["nome", "turma", "nota1", "nota2", "nota3", "nota4", "somatorio", "media", "status"]].copy()
                    df_turma.columns = ["Nome do Aluno", "Turma", "Nota 1", "Nota 2", "Nota 3", "Nota 4", "Somatório", "Média", "Status"]
                    
                    export_data, is_excel = converter_para_excel_estilizado(df_turma)
                    
                    if is_excel:
                        st.download_button(
                            label=f"📥 Baixar Excel do {turma_sel} (.xlsx)",
                            data=export_data,
                            file_name=f"boletim_{turma_sel.replace(' ', '_')}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True
                        )
                    else:
                        st.download_button(
                            label=f"📥 Baixar CSV do {turma_sel} (.csv)",
                            data=export_data,
                            file_name=f"boletim_{turma_sel.replace(' ', '_')}.csv",
                            mime="text/csv",
                            use_container_width=True
                        )
                        st.warning("⚠️ **Nota de Compatibilidade:** O pacote `openpyxl` não está instalado no seu ambiente Streamlit Cloud, por isso a planilha foi gerada automaticamente como **CSV** (totalmente editável e compatível com Excel). Para baixar o formato com design estilizado (.xlsx), inclua `openpyxl` no arquivo `requirements.txt` do seu repositório no GitHub!")
            else:
                st.warning("Nenhuma turma cadastrada no momento.")
                
        elif tipo_exportacao == "Boletim Individual de um Aluno":
            df_alunos["Identificador"] = df_alunos["nome"] + " (" + df_alunos["turma"] + ")"
            lista_nomes = sorted(df_alunos["Identificador"].unique().tolist())
            
            if lista_nomes:
                with col_exp_1:
                    aluno_sel = st.selectbox("Selecione o Aluno:", lista_nomes)
                with col_exp_2:
                    df_ind = df_alunos[df_alunos["Identificador"] == aluno_sel][["nome", "turma", "nota1", "nota2", "nota3", "nota4", "somatorio", "media", "status"]].copy()
                    df_ind.columns = ["Nome do Aluno", "Turma", "Nota 1", "Nota 2", "Nota 3", "Nota 4", "Somatório", "Média", "Status"]
                    
                    export_data, is_excel = converter_para_excel_estilizado(df_ind)
                    nome_aluno_limpo = df_ind["Nome do Aluno"].values[0].replace(" ", "_")
                    
                    if is_excel:
                        st.download_button(
                            label=f"📥 Baixar Excel de: {df_ind['Nome do Aluno'].values[0]} (.xlsx)",
                            data=export_data,
                            file_name=f"boletim_individual_{nome_aluno_limpo}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True
                        )
                    else:
                        st.download_button(
                            label=f"📥 Baixar CSV de: {df_ind['Nome do Aluno'].values[0]} (.csv)",
                            data=export_data,
                            file_name=f"boletim_individual_{nome_aluno_limpo}.csv",
                            mime="text/csv",
                            use_container_width=True
                        )
                        st.warning("⚠️ **Nota de Compatibilidade:** O pacote `openpyxl` não está instalado no seu ambiente Streamlit Cloud, por isso a planilha foi gerada automaticamente como **CSV** (totalmente editável e compatível com Excel). Para baixar o formato com design estilizado (.xlsx), inclua `openpyxl` no arquivo `requirements.txt` do seu repositório no GitHub!")
            else:
                st.warning("Nenhum aluno cadastrado no momento.")
                
        # LIMPEZA DO BANCO DE DADOS
        st.markdown("---")
        st.markdown("### ⚠️ Zona de Perigo")
        st.markdown('<div class="danger-button">', unsafe_allow_html=True)
        if st.button("🗑️ Resetar Banco de Dados (Excluir Todos os Alunos)"):
            conn = get_db_connection()
            cursor = conn.cursor()
            cursor.execute("DELETE FROM alunos")
            conn.commit()
            conn.close()
            st.warning("Todos os alunos foram excluídos do banco de dados!")
            st.rerun()
        st.markdown('</div>', unsafe_allow_html=True)
                
    else:
        st.info("Nenhum aluno cadastrado no momento. Acesse '📋 Cadastrar Aluno' para registrar os dados.")
